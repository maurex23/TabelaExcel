import numpy as np
from openpyxl import Workbook
with open('TABFER.TXT', 'r') as arq:
    a  = arq.read()
    b = open('TABFER.TXT', 'r')
    linhas = b.readlines()
    #-----------Excel
    wb = Workbook()
    dest_filename = 'Tabela.xlsx'
    aba1 = wb.active
    aba1.title = "Quantitativo"
    intervalo = 17
    for linha in range(intervalo, len(linhas)):
        valor = linhas[linha].split()
        if len(valor) == 1:
            value = str(valor)
            aba1.cell(column=1, row=linha - intervalo+1, value=value)
        if len(valor) == 2:
            for coluna in range(0, 2):
                value = str(valor[coluna])
                aba1.cell(column=coluna+1, row=linha - intervalo+1, value=value)
        if len(valor) == 3:
            for coluna in range(0, 3):
                value = str(valor[coluna])
                aba1.cell(column=coluna+1, row=linha - intervalo+1, value=value)
        if len(valor) == 4:
            for coluna in range(0, 4):
                value = str(valor[coluna])
                aba1.cell(column=coluna+1, row=linha - intervalo+1, value=value)
        if len(valor) == 5:
            for coluna in range(0, 5):
                value = str(valor[coluna])
                aba1.cell(column=coluna+1, row=linha - intervalo+1, value=value)
        if len(valor) == 6:
            for coluna in range(0, 6):
                value = str(valor[coluna])
                aba1.cell(column=coluna + 1, row=linha - intervalo+1, value=value)
        if len(valor) == 7:
            for coluna in range(0, 7):
                value = str(valor[coluna])
                aba1.cell(column=coluna + 1, row=linha - intervalo+1, value=value)
        if len(valor) == 8:
            for coluna in range(0, 8):
                value = str(valor[coluna])
                aba1.cell(column=coluna + 1, row=linha - intervalo+1, value=value)



    wb.save(filename = dest_filename)
